import os
import re
import sys
from pathlib import Path
from openpyxl import load_workbook
def log_ok(msg):   print(f"[OK] {msg}",   flush=True)
def log_info(msg): print(f"[INFO] {msg}", flush=True)
def log_warn(msg): print(f"[WARN] {msg}", flush=True)
def log_err(msg):  print(f"[ERR] {msg}",  flush=True)

SCRIPT_DIR = Path(__file__).resolve().parent
APP_BASE   = Path(os.environ.get("APP_BASE", SCRIPT_DIR.parent)).resolve()   
DOCS_DIR   = Path(os.environ.get("DOCS_DIR", APP_BASE / "docs")).resolve()

INPUT_FILE  = DOCS_DIR / "updated_tvl.xlsx"
OUTPUT_FILE = DOCS_DIR / "Weekly_Performance_updated.xlsx"
SHEET_NAME  = "ONCHAIN"
START_ROW   = 4
TVL_COL     = "O"

CELL_REF = re.compile(r'(?P<col>\$?[A-Za-z]{1,3}\$?)(?P<row>\d+)\b')

def adjust_formula(formula: str, old_row: int, new_row: int) -> str:
    def repl(m):
        if int(m.group('row')) == old_row:
            return f"{m.group('col')}{new_row}"
        return m.group(0)
    return CELL_REF.sub(repl, formula)

def cell_val(ws, col, row):
    return ws[f"{col}{row}"].value

def read_row(ws, row_idx, max_col):
    out = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=c)
        val = cell.value
        out.append((val, isinstance(val, str) and val.startswith('=')))
    return out

def write_row(ws, target_row, row_data, old_row):
    for c, (val, was_formula) in enumerate(row_data, start=1):
        if was_formula and isinstance(val, str):
            val = adjust_formula(val, old_row, target_row)
        ws.cell(row=target_row, column=c, value=val)

def main():
    if not INPUT_FILE.exists():
        log_err(f"Input file not found")
        sys.exit(1)

    wb = load_workbook(str(INPUT_FILE))
    if SHEET_NAME not in wb.sheetnames:
        log_err(f"Sheet '{SHEET_NAME}' not found")
        sys.exit(1)
    ws = wb[SHEET_NAME]

    max_row = ws.max_row
    max_col = ws.max_column

    blocks = []
    in_block = False
    start = None

    for r in range(START_ROW, max_row + 1):
        v = cell_val(ws, TVL_COL, r)
        if isinstance(v, (int, float)):
            if not in_block:
                start = r
                in_block = True
        else:
            if in_block:
                blocks.append((start, r - 1))
                in_block = False
    if in_block:
        blocks.append((start, max_row))

    for (s, e) in blocks:
        rows = []
        for orig in range(s, e + 1):
            tvl = cell_val(ws, TVL_COL, orig)
            if isinstance(tvl, (int, float)):
                data = read_row(ws, orig, max_col)
                rows.append((orig, data, tvl))

        rows.sort(key=lambda x: x[2], reverse=True)

        for idx, (orig, data, _) in enumerate(rows):
            dest = s + idx
            write_row(ws, dest, data, orig)

        log_ok(f"Sorted rows {s}–{e} by TVL")

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_FILE))
    log_ok(f"Successfully sorted by TVL")

if __name__ == "__main__":
    main()
