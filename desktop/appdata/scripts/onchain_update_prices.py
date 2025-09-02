from pathlib import Path
from dotenv import load_dotenv
import os
import sys
import time
import requests
from openpyxl import load_workbook
def log_ok(msg):   print(f"[OK] {msg}",   flush=True)
def log_info(msg): print(f"[INFO] {msg}", flush=True)
def log_warn(msg): print(f"[WARN] {msg}", flush=True)
def log_err(msg):  print(f"[ERR] {msg}",  flush=True)

SCRIPT_DIR = Path(__file__).resolve().parent
APP_BASE   = Path(os.environ.get("APP_BASE", SCRIPT_DIR.parent)).resolve()  
DOCS_DIR   = Path(os.environ.get("DOCS_DIR", APP_BASE / "docs")).resolve()

INPUT_FILE  = DOCS_DIR / "updated_file.xlsx"
OUTPUT_FILE = DOCS_DIR / "updated_prices.xlsx"

for env_path in [APP_BASE / ".env", SCRIPT_DIR / ".env", DOCS_DIR / ".env"]:
    if env_path.exists():
        load_dotenv(env_path)

API_KEY = os.getenv("API_KEY")
if not API_KEY:
    log_err("API_KEY is missing. Put it in .env (APP_BASE/.env) or set it in the environment.")
    sys.exit(1)

SHEET_NAME       = "ONCHAIN"
START_ROW        = 4
NAME_COL         = "B"
SYMBOL_COL       = "C"
PRICE_COL        = "H"
STOP_EMPTY_LIMIT = 10
REQUEST_DELAY    = 2.1

CMC_MAP_URL   = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/map"
CMC_QUOTE_URL = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/quotes/latest"
HEADERS       = {"X-CMC_PRO_API_KEY": API_KEY}

def fetch(url, params=None):
    try:
        r = requests.get(url, headers=HEADERS, params=params, timeout=20)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        log_err(f"CMC request failed for {url}: {e}")
        sys.exit(1)

log_info("Fetching CMC mapping…")
cmc_map = fetch(CMC_MAP_URL).get("data", [])
symbol_map = {}
for entry in cmc_map:
    sym = str(entry.get("symbol", "")).upper()
    if sym:
        symbol_map.setdefault(sym, []).append(entry)

if not INPUT_FILE.exists():
    log_err(f"Input file not found")
    sys.exit(1)

wb = load_workbook(str(INPUT_FILE))
if SHEET_NAME not in wb.sheetnames:
    log_err(f"Sheet '{SHEET_NAME}' not found")
    sys.exit(1)
ws = wb[SHEET_NAME]

row = START_ROW
empty_count = 0

while empty_count < STOP_EMPTY_LIMIT:
    name_val   = ws[f"{NAME_COL}{row}"].value
    symbol_val = ws[f"{SYMBOL_COL}{row}"].value

    if symbol_val is None or (isinstance(symbol_val, str) and not symbol_val.strip()):
        empty_count += 1
        row += 1
        continue

    empty_count = 0
    ticker = str(symbol_val).strip().upper()
    if ticker == "SYMBOLS":
        row += 1
        continue

    coin_name = str(name_val).strip().lower() if name_val else None

    candidates = symbol_map.get(ticker, [])
    cmc_id = None
    if not candidates:
        log_warn(f"No CMC map entry for {ticker}")
    elif len(candidates) == 1:
        cmc_id = candidates[0].get("id")
    else:
        match = next((c for c in candidates if str(c.get("name", "")).lower() == (coin_name or "")), None)
        cmc_id = (match or candidates[0]).get("id")
        if match is None:
            log_warn(f"Ambiguous ticker {ticker}, using id {cmc_id}")

    params = {"convert": "USD"}
    if cmc_id:
        params["id"] = cmc_id
        key = str(cmc_id)
    else:
        params["symbol"] = ticker
        key = ticker

    try:
        data = fetch(CMC_QUOTE_URL, params=params)
        price = data["data"][key]["quote"]["USD"]["price"]
        ws[f"{PRICE_COL}{row}"] = float(price)
        log_ok(f"{ticker}: ${price:.4f}")
    except Exception as e:
        log_err(f"Error fetching price for {ticker}: {e}")

    time.sleep(REQUEST_DELAY)
    row += 1

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUTPUT_FILE))
log_ok(f"Prices updated")
