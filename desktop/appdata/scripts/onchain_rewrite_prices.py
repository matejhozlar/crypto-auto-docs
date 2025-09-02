from openpyxl import load_workbook
import sys

def log_ok(msg):   print(f"[OK] {msg}",   flush=True)
def log_info(msg): print(f"[INFO] {msg}", flush=True)
def log_warn(msg): print(f"[WARN] {msg}", flush=True)
def log_err(msg):  print(f"[ERR] {msg}",  flush=True)

INPUT_FILE       = "../docs/Monthly_Performance_CVR.xlsx"
OUTPUT_FILE      = "../docs/updated_file.xlsx"
SHEET_NAME       = "ONCHAIN"
STOP_EMPTY_LIMIT = 10

try:
    wb = load_workbook(INPUT_FILE)
except FileNotFoundError:
    log_err(f"Could not find {INPUT_FILE}", file=sys.stderr)
    sys.exit(1)

if SHEET_NAME not in wb.sheetnames:
    log_err(f"Sheet '{SHEET_NAME}' not found in {INPUT_FILE}", file=sys.stderr)
    sys.exit(1)

ws = wb[SHEET_NAME]

empty_count = 0
row = 4

while empty_count < STOP_EMPTY_LIMIT:
    price_cell = ws[f"H{row}"]
    target_cell = ws[f"F{row}"]

    val = price_cell.value
    if isinstance(val, (int, float)):
        target_cell.value = val
        empty_count = 0
    else:
        empty_count += 1

    row += 1

wb.save(OUTPUT_FILE)
log_ok(f"Successfully rewrote prices into file: {OUTPUT_FILE}")
