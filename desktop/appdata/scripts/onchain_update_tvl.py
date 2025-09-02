import sys
import os
import time
from pathlib import Path

import requests
from openpyxl import load_workbook
def log_ok(msg):   print(f"[OK] {msg}",   flush=True)
def log_info(msg): print(f"[INFO] {msg}", flush=True)
def log_warn(msg): print(f"[WARN] {msg}", flush=True)
def log_err(msg):  print(f"[ERR] {msg}",  flush=True)

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_DOCS_DIR = (SCRIPT_DIR / ".." / "docs").resolve()  
DOCS_DIR = Path(os.environ.get("DOCS_DIR", str(DEFAULT_DOCS_DIR))).resolve()

INPUT_FILE  = DOCS_DIR / "updated_prices.xlsx"
OUTPUT_FILE = DOCS_DIR / "updated_tvl.xlsx"

SHEET_NAME       = "ONCHAIN"
START_ROW        = 4
SYMBOL_COL       = "C"
SLUG_COL         = "D"
TYPE_COL         = "E"
TVL_COL          = "O"

STOP_EMPTY_LIMIT = 10
REQUEST_DELAY    = 1.0

def fetch(url):
    try:
        r = requests.get(url, timeout=20)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        log_err(f"Failed to fetch {url}: {e}")
        sys.exit(1)

def fetch_single_protocol(slug: str) -> dict:
    return fetch(f"https://api.llama.fi/protocol/{slug}")

def fetch_all_chains() -> list:
    return fetch("https://api.llama.fi/chains")

def compute_protocol_tvl(data: dict) -> float:
    cc = data.get("currentChainTvls", {})
    deposits = sum(
        v for k, v in cc.items()
        if isinstance(v, (int, float))
           and "-" not in k
           and k not in ("staking", "borrowed")
    )
    staking  = cc.get("staking", 0) or 0
    borrowed = cc.get("borrowed", 0) or 0
    return deposits + staking + borrowed

chains = fetch_all_chains()
chain_map = {}
for c in chains:
    ts = (c.get("tokenSymbol") or "").upper()
    if ts:
        chain_map[ts] = c
    name = (c.get("name") or "").upper()
    if name:
        chain_map[name] = c

if not INPUT_FILE.exists():
    log_err(f"Input file not found")
    sys.exit(1)

wb = load_workbook(str(INPUT_FILE))
if SHEET_NAME not in wb.sheetnames:
    log_err(f"Sheet '{SHEET_NAME}' not found")
    sys.exit(1)
ws = wb[SHEET_NAME]

row = START_ROW
empty_count = 0

while empty_count < STOP_EMPTY_LIMIT:
    symbol = ws[f"{SYMBOL_COL}{row}"].value
    slug   = ws[f"{SLUG_COL}{row}"].value
    typ    = ws[f"{TYPE_COL}{row}"].value
    tvl_cell = ws[f"{TVL_COL}{row}"]

    if not symbol or (isinstance(symbol, str) and not symbol.strip()):
        empty_count += 1
        row += 1
        continue

    empty_count = 0
    symbol = str(symbol).strip()
    slug   = (slug or "").strip()
    typ    = (typ or "").strip().lower()

    if slug.lower() == "slug" or typ.lower() == "type":
        row += 1
        continue

    total_tvl = None

    if typ == "protocol" and slug:
        try:
            data = fetch_single_protocol(slug)
            total_tvl = compute_protocol_tvl(data)
            log_ok(f"Protocol {symbol:<8} ({slug}) → TVL = ${total_tvl:,.2f}")
        except Exception as e:
            log_warn(f"Protocol error for {symbol}/{slug}: {e}")

    elif typ == "chain":
        entry = chain_map.get(slug.upper()) or chain_map.get(symbol.upper())
        if entry:
            total_tvl = entry.get("tvlUsd") or entry.get("tvl") or 0
            log_ok(f"Chain    {symbol:<8} ({entry.get('name','?')}) → TVL = ${total_tvl:,.2f}")
        else:
            log_warn(f"Chain not found for slug/symbol '{slug or symbol}'")

    else:
        log_warn(f"Row {row}: unknown type '{typ}' or missing slug '{slug}'")

    if total_tvl and total_tvl > 0:
        tvl_cell.value = round(float(total_tvl), 2)
    else:
        tvl_cell.value = "N/A"

    row += 1
    time.sleep(REQUEST_DELAY)

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUTPUT_FILE))
log_ok(f"TVL update complete")
