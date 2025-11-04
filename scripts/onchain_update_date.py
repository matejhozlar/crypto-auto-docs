import os
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

def log_ok(msg):   print(f"[OK] {msg}",   flush=True)
def log_info(msg): print(f"[INFO] {msg}", flush=True)
def log_warn(msg): print(f"[WARN] {msg}", flush=True)
def log_err(msg):  print(f"[ERR] {msg}",  flush=True)

SCRIPT_DIR = Path(__file__).resolve().parent
APP_BASE   = Path(os.environ.get("APP_BASE", SCRIPT_DIR.parent)).resolve()
DOCS_DIR   = Path(os.environ.get("DOCS_DIR", APP_BASE / "docs")).resolve()

INPUT_FILE  = DOCS_DIR / "Monthly_Performance_CVR_latest.xlsx"
OUTPUT_FILE = DOCS_DIR / "Monthly_Performance_CVR_latest.xlsx"
SHEET_NAME  = "ONCHAIN"
DATE_CELL   = "B1"

def main():
    if not INPUT_FILE.exists():
        log_err(f"Input file not found: {INPUT_FILE}")
        sys.exit(1)

    try:
        wb = load_workbook(str(INPUT_FILE))
    except Exception as e:
        log_err(f"Failed to open workbook: {e}")
        sys.exit(1)

    if SHEET_NAME not in wb.sheetnames:
        log_err(f"Sheet '{SHEET_NAME}' not found")
        sys.exit(1)

    ws = wb[SHEET_NAME]

    current_date = datetime.now().strftime("%d.%m.%Y")
    
    ws[DATE_CELL].value = current_date
    
    log_info(f"Updated {DATE_CELL} to: {current_date}")

    try:
        OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(OUTPUT_FILE))
        log_ok(f"Date update completed successfully.")
    except Exception as e:
        log_err(f"Failed to save workbook: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()