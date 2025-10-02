import os
import re
import sys
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import coordinate_to_tuple

def log_ok(msg):   print(f"[OK] {msg}",   flush=True)
def log_info(msg): print(f"[INFO] {msg}", flush=True)
def log_warn(msg): print(f"[WARN] {msg}", flush=True)
def log_err(msg):  print(f"[ERR] {msg}",  flush=True)

SCRIPT_DIR = Path(__file__).resolve().parent
APP_BASE   = Path(os.environ.get("APP_BASE", SCRIPT_DIR.parent)).resolve()
DOCS_DIR   = Path(os.environ.get("DOCS_DIR", APP_BASE / "docs")).resolve()

INPUT_FILE   = DOCS_DIR / "sorted_tvl.xlsx"
OUTPUT_FILE  = DOCS_DIR / "validated.xlsx"
SHEET_NAME   = "ONCHAIN"

START_ROW          = 4
RATIO_COL_LETTER   = "S"   
MARK_COL_LETTER    = "B"   
MAX_EMPTY_STREAK   = 10
RATIO_THRESHOLD    = 100.0

RED_FILL = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")

SKIP_TOKENS = {"fd mcap / tvl ratio"}

NUM_TOKEN = re.compile(
    r"^\s*([0-9]{1,3}(?:[ ,\u00A0\u202F]?[0-9]{3})*(?:[.,][0-9]+)?|[0-9]+(?:[.,][0-9]+)?)\s*[xX×]?\s*$"
)

DEBUG_PARSE = False

def parse_ratio(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            return float(val)
        except Exception:
            return None
    if isinstance(val, str):
        s = val.strip()
        if s.lower() in SKIP_TOKENS:
            return None
        s = s.replace("\u00A0", " ").replace("\u202F", " ")
        m = NUM_TOKEN.match(s)
        if not m:
            if DEBUG_PARSE:
                log_info(f"[parse] no match: {repr(val)}")
            return None
        num = m.group(1).replace(" ", "")
        if "," in num and "." in num:
            num = num.replace(",", "")
        elif "," in num:
            num = num.replace(",", ".")
        try:
            return float(num)
        except Exception:
            if DEBUG_PARSE:
                log_info(f"[parse] float fail after normalize: {repr(num)} from {repr(val)}")
            return None
    return None

def is_empty_cell(val):
    return (val is None) or (isinstance(val, str) and val.strip() == "")

def cell_addr(col_letter: str, row: int) -> str:
    return f"{col_letter}{row}"

def backup_file(src_path: Path):
    bak_path = src_path.with_suffix(src_path.suffix + ".bak")
    try:
        shutil.copyfile(src_path, bak_path)
        log_info(f"Backup created: {bak_path}")
    except Exception as e:
        log_warn(f"Could not create backup ({e}). Continuing without backup.")

def get_last_row(ws):
    try:
        dim = ws.calculate_dimension()  
        if ":" in dim:
            _, br = dim.split(":")
            last_row, _ = coordinate_to_tuple(br)
            if 0 < last_row < 1_000_000:
                return last_row
    except Exception:
        pass
    mr = getattr(ws, "max_row", None)
    if isinstance(mr, int) and 0 < mr < 1_000_000:
        return mr
    return START_ROW + 10000

def main():
    if not INPUT_FILE.exists():
        log_err(f"Input file not found: {INPUT_FILE}")
        sys.exit(1)

    try:
        wb_values = load_workbook(str(INPUT_FILE), data_only=True, read_only=True)
        wb_raw    = load_workbook(str(INPUT_FILE), data_only=False, read_only=True)
    except Exception as e:
        log_err(f"Failed to open workbook for reading: {e}")
        sys.exit(1)

    if SHEET_NAME not in wb_values.sheetnames or SHEET_NAME not in wb_raw.sheetnames:
        log_err(f"Sheet '{SHEET_NAME}' not found in {INPUT_FILE.name}")
        sys.exit(1)

    ws_values = wb_values[SHEET_NAME]
    ws_raw    = wb_raw[SHEET_NAME]

    def read_ratio_cell(col_letter, row):
        v = ws_values[cell_addr(col_letter, row)].value
        if v is not None and not is_empty_cell(v):
            return v
        raw_cell = ws_raw[cell_addr(col_letter, row)]
        if getattr(raw_cell, "data_type", None) == "f":
            cached = getattr(raw_cell, "cached_value", None)
            if cached is None:
                cached = getattr(raw_cell, "_value", None)
            return cached
        return raw_cell.value

    last_row = get_last_row(ws_values)
    log_info(f"Scanning {SHEET_NAME}!{RATIO_COL_LETTER}{START_ROW}:{RATIO_COL_LETTER}{last_row}")

    rows_to_flag = []
    checked = 0
    empty_streak = 0

    for row in range(START_ROW, last_row + 1):
        try:
            ratio_val = read_ratio_cell(RATIO_COL_LETTER, row)
        except Exception:
            break

        if is_empty_cell(ratio_val):
            empty_streak += 1
            if empty_streak >= MAX_EMPTY_STREAK and row > START_ROW + 50:
                log_info(f"Many empties encountered by row {row}. Stopping early as a safety.")
                break
            continue

        empty_streak = 0
        ratio = parse_ratio(ratio_val)
        if ratio is not None:
            checked += 1
            if ratio > RATIO_THRESHOLD:
                rows_to_flag.append(row)
        elif DEBUG_PARSE:
            log_info(f"[parse] unparsable at row {row}: {repr(ratio_val)}")

    try:
        wb_values.close()
        wb_raw.close()
    except Exception:
        pass

    backup_file(INPUT_FILE)

    try:
        wb = load_workbook(str(INPUT_FILE), data_only=False, read_only=False, keep_vba=False)
    except Exception as e:
        log_err(f"Failed to reopen workbook for writing: {e}")
        sys.exit(1)

    if SHEET_NAME not in wb.sheetnames:
        log_err(f"Sheet '{SHEET_NAME}' not found.")
        sys.exit(1)

    ws = wb[SHEET_NAME]

    flagged = 0
    for r in rows_to_flag:
        try:
            mark_cell = ws[cell_addr(MARK_COL_LETTER, r)]
            mark_cell.fill = RED_FILL
            flagged += 1
        except Exception as e:
            log_warn(f"Could not color row {r}: {e}")

    try:
        OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(OUTPUT_FILE))
        wb.close()
    except Exception as e:
        log_err(f"Failed to save output workbook: {e}")
        sys.exit(1)

    log_warn(f"Flagged {flagged} rows.\n")
    log_ok(f"Validation completed successfully.")
    log_ok(f"Wrote: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
