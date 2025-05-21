#!/usr/bin/env python3
from dotenv import load_dotenv
import os
import sys
import requests
from openpyxl import load_workbook
import time

load_dotenv()

API_KEY          = os.getenv("API_KEY")
INPUT_FILE       = "../docs/updated_file.xlsx"
OUTPUT_FILE      = "../docs/updated_prices.xlsx"
SHEET_NAME       = "ONCHAIN"
START_ROW        = 4
NAME_COL         = 'B'   
SYMBOL_COL       = 'C'   
PRICE_COL        = 'H'   
STOP_EMPTY_LIMIT = 10
REQUEST_DELAY    = 2.1

CMC_MAP_URL      = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/map"
CMC_QUOTE_URL    = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/quotes/latest"
HEADERS          = {"X-CMC_PRO_API_KEY": API_KEY}

def fetch(url, params=None):
    r = requests.get(url, headers=HEADERS, params=params)
    if r.status_code != 200:
        sys.exit(f"❌ ERROR: CMC API {url} returned HTTP {r.status_code}")
    return r.json()

print("Fetching CMC mapping...", file=sys.stderr)
cmc_map = fetch(CMC_MAP_URL)["data"]
symbol_map = {}
for entry in cmc_map:
    sym = entry["symbol"].upper()
    symbol_map.setdefault(sym, []).append(entry)

if not os.path.exists(INPUT_FILE):
    print(f"❌ ERROR: Input file not found: {INPUT_FILE}", file=sys.stderr)
    sys.exit(1)

wb = load_workbook(INPUT_FILE)
if SHEET_NAME not in wb.sheetnames:
    print(f"❌ ERROR: Sheet '{SHEET_NAME}' not found in {INPUT_FILE}", file=sys.stderr)
    sys.exit(1)
ws = wb[SHEET_NAME]

row = START_ROW
empty_count = 0

while empty_count < STOP_EMPTY_LIMIT:
    name   = ws[f"{NAME_COL}{row}"].value
    symbol = ws[f"{SYMBOL_COL}{row}"].value

    if str(symbol).strip().upper() == "SYMBOLS":
        empty_count = 0
        row += 1
        continue

    if symbol is None:
        empty_count += 1
        row += 1
        continue
    empty_count = 0

    ticker = str(symbol).strip().upper()
    coin_name = str(name).strip().lower() if name else None

    candidates = symbol_map.get(ticker, [])
    if not candidates:
        print(f"⚠️  WARNING: No CMC map entry for {ticker}", file=sys.stderr)
        cmc_id = None
    elif len(candidates) == 1:
        cmc_id = candidates[0]["id"]
    else:
        match = next(
            (c for c in candidates if c["name"].lower() == coin_name),
            None
        )
        cmc_id = (match or candidates[0])["id"]
        if match is None:
            print(f"⚠️  NOTE: Ambiguous ticker {ticker}, using id {cmc_id}", file=sys.stderr)

    if cmc_id:
        params = {"id": cmc_id, "convert": "USD"}
    else:
        params = {"symbol": ticker, "convert": "USD"}

    try:
        data = fetch(CMC_QUOTE_URL, params=params)
        key = str(cmc_id) if cmc_id else ticker
        price = data["data"][key]["quote"]["USD"]["price"]
        ws[f"{PRICE_COL}{row}"] = price
        print(f"✅ {ticker}: ${price:.4f}")
    except Exception as e:
        print(f"❌ Error fetching price for {ticker}: {e}", file=sys.stderr)

    time.sleep(REQUEST_DELAY)
    row += 1

wb.save(OUTPUT_FILE)
print(f"✅ Prices updated in {OUTPUT_FILE}")
