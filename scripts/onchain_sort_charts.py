import os
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

def log_ok(msg):   print(f"[OK] {msg}",   flush=True)
def log_info(msg): print(f"[INFO] {msg}", flush=True)
def log_warn(msg): print(f"[WARN] {msg}", flush=True)
def log_err(msg):  print(f"[ERR] {msg}",  flush=True)

SCRIPT_DIR = Path(__file__).resolve().parent
APP_BASE   = Path(os.environ.get("APP_BASE", SCRIPT_DIR.parent)).resolve()
DOCS_DIR   = Path(os.environ.get("DOCS_DIR", APP_BASE / "docs")).resolve()

INPUT_FILE   = DOCS_DIR / "validated.xlsx"
OUTPUT_FILE  = DOCS_DIR / "Monthly_Performance_CVR.xlsx"
SHEET_NAME   = "CHARTS"
START_CELL   = "C4"
KEY_COL      = "C"
DESCENDING   = False

TRY_EXCEL_RECALC = True
REQUIRE_NUMERIC_KEYS = True

CELL_OR_XSHEET = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_]+)\!)?(?P<col>\$?[A-Za-z]{1,3}\$?)(?P<row>\d+)\b"
)
SINGLE_REF_FORMULA = re.compile(
    r"^\s*=\s*(?:(?P<sheet>'[^']+'|[A-Za-z0-9_]+)\!)?(?P<col>\$?[A-Za-z]{1,3})\$?(?P<row>\d+)\s*$"
)

def col_row_from_a1(a1: str):
    m = re.match(r"^\$?([A-Za-z]{1,3})\$(\d+)$|^\$?([A-Za-z]{1,3})(\d+)$", a1)
    if not m:
        raise ValueError(f"Bad A1: {a1}")
    col = m.group(1) or m.group(3)
    row = int(m.group(2) or m.group(4))
    return col, row

def get_start_coords():
    return col_row_from_a1(START_CELL)

def is_blank(val):
    return val is None or (isinstance(val, str) and val.strip() == "")

def try_excel_recalc(path: Path):
    if os.name != "nt":
        return False
    try:
        import win32com.client
    except Exception as e:
        log_warn(f"Excel recalculation unavailable (pywin32 missing?): {e}")
        return False

    try:
        log_info("Recalculating workbook in Excel (Windows)...")
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.ScreenUpdating = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=False)
            excel.CalculateFullRebuild()
            wb.Save()
            wb.Close(SaveChanges=True)
        finally:
            excel.Quit()
        log_ok("Excel recalculation completed successfully.")
        return True
    except Exception as e:
        log_warn(f"Excel recalculation failed ({e}).")
        return False

def load_pair(path: Path):
    wb_w = load_workbook(str(path), data_only=False)
    wb_v = load_workbook(str(path), data_only=True)
    return wb_w, wb_v

def resolve_value(wb_w, wb_v, sheet: str, col: str, row: int, max_depth: int = 5):
    cur_sheet, cur_col, cur_row = sheet, col, row
    for _ in range(max_depth + 1):
        try:
            v = wb_v[cur_sheet][f"{cur_col}{cur_row}"].value
        except KeyError:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        f = wb_w[cur_sheet][f"{cur_col}{cur_row}"].value
        if isinstance(f, str) and f.startswith("="):
            m = SINGLE_REF_FORMULA.match(f)
            if m:
                s = m.group("sheet")
                cur_sheet = (s.strip("'") if s else cur_sheet)
                cur_col   = m.group("col")
                cur_row   = int(m.group("row"))
                continue
        return None
    return None

def read_row(ws, row_idx, max_col):
    out = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=c)
        val = cell.value
        out.append((val, isinstance(val, str) and val.startswith('=')))
    return out

def adjust_formula(formula: str, old_row: int, new_row: int, current_sheet: str) -> str:
    def repl(m):
        sheet = m.group('sheet')
        col = m.group('col')
        row = int(m.group('row'))
        if sheet:
            if sheet.strip("'") != current_sheet:
                return m.group(0)  
        if row == old_row:
            return (f"{sheet + '!' if sheet else ''}{col}{new_row}")
        return m.group(0)
    return CELL_OR_XSHEET.sub(repl, formula)

def write_row(ws, target_row, row_data, old_row, current_sheet):
    for c, (val, was_formula) in enumerate(row_data, start=1):
        if was_formula and isinstance(val, str):
            val = adjust_formula(val, old_row, target_row, current_sheet)
        ws.cell(row=target_row, column=c, value=val)

def find_blocks(ws, key_col, start_row, max_row):
    blocks = []
    in_block = False
    start = None
    for r in range(start_row, max_row + 1):
        v = ws[f"{key_col}{r}"].value
        if not is_blank(v):
            if not in_block:
                start = r
                in_block = True
        else:
            if in_block:
                blocks.append((start, r - 1))
                in_block = False
    if in_block:
        blocks.append((start, max_row))
    return blocks

def main():
    if not INPUT_FILE.exists():
        log_err("Input file not found")
        sys.exit(1)

    wb_w, wb_v = load_pair(INPUT_FILE)
    if SHEET_NAME not in wb_w.sheetnames:
        log_err(f"Sheet '{SHEET_NAME}' not found")
        sys.exit(1)
    ws_w = wb_w[SHEET_NAME]
    ws_v = wb_v[SHEET_NAME]

    start_col, start_row = get_start_coords()
    max_row = ws_w.max_row
    max_col = ws_w.max_column

    blocks = find_blocks(ws_w, KEY_COL, start_row, max_row)
    if not blocks:
        log_warn("No table blocks detected to sort.")
        OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
        wb_w.save(str(OUTPUT_FILE))
        return

    any_numeric = False
    for (s, e) in blocks:
        for r in range(s, e + 1):
            v = ws_v[f"{KEY_COL}{r}"].value
            if isinstance(v, (int, float)):
                any_numeric = True
                break

            # NEW: if there is no cached value (Linux), try resolving single-cell references
            rv = resolve_value(wb_w, wb_v, SHEET_NAME, KEY_COL, r)
            if isinstance(rv, (int, float)):
                any_numeric = True
                break

        if any_numeric:
            break

    # Only attempt Excel COM recalculation on Windows if we still couldn't resolve anything
    if not any_numeric and TRY_EXCEL_RECALC and os.name == "nt":
        if try_excel_recalc(INPUT_FILE):
            wb_w, wb_v = load_pair(INPUT_FILE)
            ws_w = wb_w[SHEET_NAME]
            ws_v = wb_v[SHEET_NAME]

            # Re-run scan after recalculation
            for (s, e) in blocks:
                for r in range(s, e + 1):
                    v = ws_v[f"{KEY_COL}{r}"].value
                    if isinstance(v, (int, float)):
                        any_numeric = True
                        break
                    rv = resolve_value(wb_w, wb_v, SHEET_NAME, KEY_COL, r)
                    if isinstance(rv, (int, float)):
                        any_numeric = True
                        break
                if any_numeric:
                    break

    if not any_numeric and REQUIRE_NUMERIC_KEYS:
        log_err(
            "Cannot sort: no cached numeric values found for key column "
            "(and no resolvable single-cell references)."
        )
        sys.exit(2)


    for (s, e) in blocks:
        start_sort_row = s if s == start_row else (s + 1)

        rows = []
        for orig in range(start_sort_row, e + 1):
            key_val = ws_v[f"{KEY_COL}{orig}"].value
            if not isinstance(key_val, (int, float)):
                key_val = resolve_value(wb_w, wb_v, SHEET_NAME, KEY_COL, orig)

            sort_key = (
                float(key_val) if isinstance(key_val, (int, float))
                else (float("-inf") if DESCENDING else float("inf"))
            )
            data = read_row(ws_w, orig, max_col)
            rows.append((orig, data, sort_key))

        rows.sort(key=lambda x: x[2], reverse=DESCENDING)

        dest = start_sort_row
        for orig, data, _ in rows:
            write_row(ws_w, dest, data, orig, SHEET_NAME)
            dest += 1

        kept = " (kept header at top)" if start_sort_row == s + 1 else ""
        log_ok(f"Sorted rows {start_sort_row}–{e} in block starting at {s}{kept}.")


    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb_w.save(str(OUTPUT_FILE))
    log_ok(f"CHARTS table sorting completed successfully -> {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
