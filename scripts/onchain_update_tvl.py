import sys
import time
import requests
from openpyxl import load_workbook

INPUT_FILE       = "../docs/updated_prices.xlsx"
OUTPUT_FILE      = "../docs/updated_tvl.xlsx"
SHEET_NAME       = "ONCHAIN"

START_ROW        = 4
SYMBOL_COL       = "C"   
SLUG_COL         = "D"   
TYPE_COL         = "E"   
TVL_COL          = "O"   

STOP_EMPTY_LIMIT = 10
REQUEST_DELAY    = 1.0   

def fetch(url):
    r = requests.get(url)
    if r.status_code != 200:
        sys.exit(f"❌ Failed to fetch {url} (HTTP {r.status_code})")
    return r.json()

def fetch_single_protocol(slug: str) -> dict:
    return fetch(f"https://api.llama.fi/protocol/{slug}")

def fetch_all_chains() -> list:
    return fetch("https://api.llama.fi/chains")

def compute_protocol_tvl(data: dict) -> float:
    cc = data.get("currentChainTvls", {})
    deposits = sum(
        v for k, v in cc.items()
        if isinstance(v, (int, float))
           and "-" not in k
           and k not in ("staking", "borrowed")
    )
    staking  = cc.get("staking", 0) or 0
    borrowed = cc.get("borrowed", 0) or 0
    return deposits + staking + borrowed


chains = fetch_all_chains()
chain_map = {}
for c in chains:
    ts = (c.get("tokenSymbol") or "").upper()
    if ts:
        chain_map[ts] = c
    name = c.get("name","").upper()
    if name:
        chain_map[name] = c


wb = load_workbook(INPUT_FILE)
if SHEET_NAME not in wb.sheetnames:
    sys.exit(f"❌ Sheet '{SHEET_NAME}' not found in {INPUT_FILE}")
ws = wb[SHEET_NAME]

row = START_ROW
empty_count = 0

while empty_count < STOP_EMPTY_LIMIT:
    symbol = ws[f"{SYMBOL_COL}{row}"].value
    slug   = ws[f"{SLUG_COL}{row}"].value
    typ    = ws[f"{TYPE_COL}{row}"].value
    tvl_cell = ws[f"{TVL_COL}{row}"]

    if not symbol or (isinstance(symbol, str) and not symbol.strip()):
        empty_count += 1
        row += 1
        continue

    empty_count = 0
    symbol = str(symbol).strip()
    slug   = (slug or "").strip()
    typ    = (typ or "").strip().lower()

    if slug.lower() == "slug" or typ.lower() == "type":
        row += 1
        continue

    total_tvl = None

    if typ == "protocol" and slug:
        try:
            data = fetch_single_protocol(slug)
            total_tvl = compute_protocol_tvl(data)
            print(f"✅ Protocol {symbol:<8} ({slug}) → TVL = ${total_tvl:,.2f}")
        except Exception as e:
            print(f"⚠️ Protocol error for {symbol}/{slug}: {e}")

    elif typ == "chain":
        entry = chain_map.get(slug.upper()) or chain_map.get(symbol.upper())
        if entry:
            total_tvl = entry.get("tvlUsd") or entry.get("tvl") or 0
            print(f"✅ Chain    {symbol:<8} ({entry['name']}) → TVL = ${total_tvl:,.2f}")
        else:
            print(f"⚠️ Chain not found for slug/symbol '{slug or symbol}'")

    else:
        print(f"⚠️ Row {row}: unknown type '{typ}' or missing slug '{slug}'")

    if total_tvl and total_tvl > 0:
        tvl_cell.value = round(total_tvl, 2)
    else:
        tvl_cell.value = "N/A"

    row += 1
    time.sleep(REQUEST_DELAY)


wb.save(OUTPUT_FILE)
print("✅ TVL update complete.")
