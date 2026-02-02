from __future__ import annotations

import os
import time
import argparse
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from typing import Optional
def log_ok(msg):   print(f"[OK] {msg}",   flush=True)
def log_info(msg): print(f"[INFO] {msg}", flush=True)
def log_warn(msg): print(f"[WARN] {msg}", flush=True)
def log_err(msg):  print(f"[ERR] {msg}",  flush=True)

SCRIPT_DIR = Path(__file__).resolve().parent              
ROOT_DIR   = SCRIPT_DIR.parent                            
DOCS_DIR   = ROOT_DIR / "docs"                            

load_dotenv(ROOT_DIR / ".env")
load_dotenv(SCRIPT_DIR / ".env")  

API_KEY = os.getenv("API_KEY")

CMC_ID_OVERRIDES = {
    "LIGHT": 37986,
}

CMC_URL  = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/quotes/latest"
HEADERS  = {"X-CMC_PRO_API_KEY": API_KEY}

SHEET_NAME              = "PERFORMANCE_TABLE"
START_ROW               = 2
SYMBOL_COL              = "C"
PRICE_COL               = "E"
UPDATE_NOTE_COL         = "G"
UPDATE_NOTE_EMPTY_BELOW = 5
STOP_EMPTY_LIMIT        = 10
REQUEST_DELAY           = 2.1 

YELLOW_RGB = "FFFF00"

def _is_empty_value(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _ordinal(n: int) -> str:
    if 11 <= (n % 100) <= 13:
        suf = "th"
    else:
        suf = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suf}"


def _format_update_date_struct(t_struct: time.struct_time | None = None) -> str:
    if t_struct is None:
        try:
            if "TZ" in os.environ and hasattr(time, "tzset"):
                time.tzset()
        except Exception:
            pass
        t_struct = time.localtime()

    month = time.strftime("%B", t_struct)
    day = t_struct.tm_mday
    year = t_struct.tm_year
    return f"{month} {_ordinal(day)}, {year}"


def _find_or_insertion_row_for_update_note(ws, col: str = UPDATE_NOTE_COL, empties_below: int = UPDATE_NOTE_EMPTY_BELOW) -> int:
    maxr = ws.max_row + 10

    for r in range(1, maxr + 1):
        v = ws[f"{col}{r}"].value
        if isinstance(v, str) and v.strip().lower().startswith("source:"):
            ok = True
            for k in range(1, empties_below + 1):
                vv = ws[f"{col}{r + k}"].value
                if not _is_empty_value(vv):
                    ok = False
                    break
            if ok:
                return r

    for r in range(1, maxr + 1):
        if _is_empty_value(ws[f"{col}{r}"].value):
            ok = True
            for k in range(1, empties_below + 1):
                if not _is_empty_value(ws[f"{col}{r + k}"].value):
                    ok = False
                    break
            if ok:
                return r

    return ws.max_row + 1


def update_last_update_note(ws) -> None:
    row = _find_or_insertion_row_for_update_note(ws, UPDATE_NOTE_COL, UPDATE_NOTE_EMPTY_BELOW)
    note = f"Source: altFINS (Last update: {_format_update_date_struct()})"
    ws[f"{UPDATE_NOTE_COL}{row}"].value = note

def _color_hex6(fg) -> Optional[str]:
    if not fg:
        return None

    for attr in ("rgb", "value"):
        v = getattr(fg, attr, None)
        if v:
            try:
                s = str(v).strip().upper()
            except Exception:
                continue
            if len(s) == 8:
                s = s[-6:]
            if len(s) == 6:
                return s

    idx = getattr(fg, "indexed", None)
    if isinstance(idx, int) and idx == 6:
        return YELLOW_RGB
    
    return None

def is_yellow(cell) -> bool:
    fill = getattr(cell, "fill", None)
    if not fill or getattr(fill, "fill_type", None) != "solid":
        return False

    fg = getattr(fill, "fgColor", None)
    hex6 = _color_hex6(fg)
    return hex6 == YELLOW_RGB

def smart_round(price: float) -> float:
    if price >= 0.01:
        return round(price, 2)
    decimals = 4
    while round(price, decimals) == 0 and decimals < 12:
        decimals += 1
    return round(price, decimals)

def fetch_cmc_price(symbol: str) -> float | None:
    symbol = symbol.strip().upper()

    params = {"convert": "USD"}
    cmc_id = CMC_ID_OVERRIDES.get(symbol)

    if cmc_id is not None:
        params["id"] = str(cmc_id)
    else:
        params["symbol"] = symbol

    r = requests.get(CMC_URL, headers=HEADERS, params=params, timeout=20)

    # If the HTTP response itself failed, log it.
    if r.status_code != 200:
        log_warn(f"CMC HTTP {r.status_code} for {symbol}: {r.text[:200]}")
        return None

    data = r.json() if r.content else {}

    # If CMC returned an API-level error, log it.
    status = data.get("status") or {}
    if status.get("error_code", 0) != 0:
        log_warn(f"CMC API error for {symbol}: {status.get('error_message')} (code {status.get('error_code')})")
        return None

    payload = data.get("data") or {}

    try:
        if cmc_id is not None:
            # Most common shape: data is keyed by the id you passed in
            item = payload.get(str(cmc_id))
            if item and "quote" in item:
                return item["quote"]["USD"]["price"]

            # Fallback: sometimes providers return dict with a single entry
            if isinstance(payload, dict) and len(payload) == 1:
                only_item = next(iter(payload.values()))
                return only_item["quote"]["USD"]["price"]

            log_warn(f"CMC returned no data for id={cmc_id} (symbol {symbol}). Keys: {list(payload)[:5]}")
            return None
        else:
            return payload[symbol]["quote"]["USD"]["price"]
    except Exception as e:
        log_warn(f"Unexpected CMC payload for {symbol}: {e}. Top-level keys: {list(data.keys())}")
        return None

def run(input_path: Path, output_path: Path) -> None:
    if not API_KEY:
        raise SystemExit("API_KEY is missing. Put it in .env at repo root or scripts/.")

    if not input_path.exists():
        raise SystemExit(f"Input file not found")

    wb = load_workbook(str(input_path))
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' not found")
    ws = wb[SHEET_NAME]

    row = START_ROW
    empty_count = 0

    while empty_count < STOP_EMPTY_LIMIT:
        symbol_cell = ws[f"{SYMBOL_COL}{row}"]
        price_cell  = ws[f"{PRICE_COL}{row}"]
        symbol      = symbol_cell.value

        if symbol is None:
            empty_count += 1
        else:
            ticker = str(symbol).strip().upper()
            if ticker == "TICKER":
                row += 1
                continue

            if not is_yellow(price_cell):
                log_warn(f"Skipping row {row} (not yellow)")
            else:
                empty_count = 0
                try:
                    price = fetch_cmc_price(ticker)
                    if price is not None:
                        rounded = smart_round(price)
                        price_cell.value = rounded
                        log_ok(f"{ticker}: ${rounded} successfully imported")
                    else:
                        log_warn(f"Symbol {ticker} not found in CMC response.")
                except Exception as e:
                    log_err(f"Error fetching {ticker}: {e}")

                time.sleep(REQUEST_DELAY)

        row += 1
    update_last_update_note(ws)

    wb.save(str(output_path))
    log_ok(f"Successfully updated prices")

def parse_args():
    p = argparse.ArgumentParser(description="Update yellow prices in PERFORMANCE_TABLE.")
    p.add_argument("--input",  type=Path, default=DOCS_DIR / "Weekly_Performance_PORTFOLIO.xlsx",
                   help="Path to input Weekly_Performance_PORTFOLIO.xlsx")
    p.add_argument("--output", type=Path, default=DOCS_DIR / "Weekly_Performance_PORTFOLIO_latest.xlsx",
                   help="Path to output workbook")
    return p.parse_args()

if __name__ == "__main__":
    args = parse_args()
    run(args.input, args.output)
