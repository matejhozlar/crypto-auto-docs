from pathlib import Path
from openpyxl import load_workbook
import os
import sys
def log_ok(msg):   print(f"[OK] {msg}",   flush=True)
def log_info(msg): print(f"[INFO] {msg}", flush=True)
def log_warn(msg): print(f"[WARN] {msg}", flush=True)
def log_err(msg):  print(f"[ERR] {msg}",  flush=True)

SCRIPT_DIR = Path(__file__).resolve().parent
APP_BASE   = Path(os.environ.get("APP_BASE", SCRIPT_DIR.parent)).resolve()    
DOCS_DIR   = Path(os.environ.get("DOCS_DIR", APP_BASE / "docs")).resolve()

INPUT_FILE  = DOCS_DIR / "Monthly_Performance_CVR.xlsx"
OUTPUT_FILE = DOCS_DIR / "updated_file.xlsx"
SHEET_NAME  = "ONCHAIN"
STOP_EMPTY_LIMIT = 10

if not INPUT_FILE.exists():
    log_err(f"File not found")
    sys.exit(1)

try:
    wb = load_workbook(str(INPUT_FILE))
except Exception as e:
    log_err(f"Failed to open workbook: {e}")
    sys.exit(1)

if SHEET_NAME not in wb.sheetnames:
    log_err(f"Sheet '{SHEET_NAME}' not found")
    sys.exit(1)

ws = wb[SHEET_NAME]

empty_count = 0
row = 4

while empty_count < STOP_EMPTY_LIMIT:
    price_cell  = ws[f"H{row}"]
    target_cell = ws[f"F{row}"]

    val = price_cell.value
    if isinstance(val, (int, float)):
        target_cell.value = val
        empty_count = 0
    else:
        empty_count += 1

    row += 1

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUTPUT_FILE))
log_ok(f"Successfully rewrote prices")
